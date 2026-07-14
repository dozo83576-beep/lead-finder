import csv
import io
import logging
import re
from concurrent.futures import ThreadPoolExecutor
from dataclasses import dataclass, field
from html.parser import HTMLParser
from typing import Any
from urllib.parse import urlparse

import requests
from openpyxl import Workbook, load_workbook


USER_AGENT = "lead-finder/2.0 (local client research tool)"
SOCIAL_HOSTS = ("vk.com", "t.me", "telegram.me", "instagram.com", "facebook.com", "wa.me")
OVERPASS_ENDPOINTS = (
    "https://maps.mail.ru/osm/tools/overpass/api/interpreter",
    "https://overpass-api.de/api/interpreter",
)
PRESETS: dict[str, tuple[tuple[str, str], ...]] = {
    "Ремонт и отделка": (
        ("craft", "builder|painter|tiler|plasterer"),
        ("office", "construction_company"),
    ),
    "Сантехники": (("craft", "plumber"),),
    "Электрики": (("craft", "electrician"),),
    "Окна и двери": (
        ("craft", "window_construction|door_construction"),
        ("shop", "windows|doors"),
    ),
    "Автосервисы": (("shop", "car_repair"),),
    "Салоны красоты": (("shop", "beauty|hairdresser"),),
    "Стоматологии": (("amenity", "dentist"), ("healthcare", "dentist")),
    "Юристы": (("office", "lawyer"),),
    "Бухгалтерия": (("office", "accountant|tax_advisor"),),
    "Клининг": (("craft", "cleaning"),),
}
STATUSES = ("Новый", "Связался", "Ответил", "Клиент", "Не подходит")
INPUT_FIELDS = (
    "name",
    "category",
    "city",
    "address",
    "phone",
    "email",
    "social",
    "website",
    "source",
    "source_url",
)
EXPORT_FIELDS = (
    "score",
    "priority",
    "reasons",
    "name",
    "category",
    "city",
    "address",
    "phone",
    "email",
    "social",
    "website",
    "status",
    "note",
    "source",
    "source_url",
)


@dataclass
class WebsiteAudit:
    state: str
    normalized_url: str = ""
    https: bool = False
    status: int | None = None
    mobile_viewport: bool | None = None
    title_present: bool | None = None
    description_present: bool | None = None
    mobile_score: int | None = None
    error: str = ""


@dataclass
class Lead:
    name: str
    lead_key: str = ""
    category: str = ""
    city: str = ""
    address: str = ""
    phone: str = ""
    email: str = ""
    social: str = ""
    website: str = ""
    source: str = ""
    source_url: str = ""
    score: int = 0
    reasons: list[str] = field(default_factory=list)
    status: str = "Новый"
    note: str = ""
    audit: WebsiteAudit | None = None

    @property
    def priority(self) -> str:
        if self.score >= 70:
            return "Высокий"
        if self.score >= 45:
            return "Средний"
        return "Низкий"


class _SignalsParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.in_title = False
        self.title_text: list[str] = []
        self.mobile_viewport = False
        self.description_present = False

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        attributes = {key.lower(): (value or "") for key, value in attrs}
        if tag.lower() == "title":
            self.in_title = True
        if tag.lower() == "meta":
            name = attributes.get("name", "").lower()
            content = attributes.get("content", "").strip()
            if name == "viewport" and content:
                self.mobile_viewport = True
            if name == "description" and content:
                self.description_present = True

    def handle_endtag(self, tag: str) -> None:
        if tag.lower() == "title":
            self.in_title = False

    def handle_data(self, data: str) -> None:
        if self.in_title:
            self.title_text.append(data)


def normalize_website(value: str) -> WebsiteAudit:
    raw = (value or "").strip()
    if not raw:
        return WebsiteAudit(state="missing")

    url = raw if "://" in raw else f"https://{raw}"
    parsed = urlparse(url)
    host = parsed.netloc.lower().split(":", 1)[0]
    if parsed.scheme not in ("http", "https") or not host or "." not in host or " " in host:
        return WebsiteAudit(state="broken", normalized_url=raw, error="Некорректный адрес")

    host = host.removeprefix("www.")
    if any(host == social or host.endswith(f".{social}") for social in SOCIAL_HOSTS):
        return WebsiteAudit(state="social", normalized_url=url, https=parsed.scheme == "https")

    return WebsiteAudit(state="unknown", normalized_url=url, https=parsed.scheme == "https")


def _parse_html_signals(html: str) -> tuple[bool, bool, bool]:
    parser = _SignalsParser()
    parser.feed(html or "")
    title_present = bool("".join(parser.title_text).strip())
    return parser.mobile_viewport, title_present, parser.description_present


def apply_pagespeed_result(audit: WebsiteAudit, data: dict[str, Any]) -> WebsiteAudit:
    raw_score = data.get("lighthouseResult", {}).get("categories", {}).get("performance", {}).get("score")
    if raw_score is not None:
        audit.mobile_score = round(float(raw_score) * 100)
    return audit


def _check_pagespeed(audit: WebsiteAudit, api_key: str, session: requests.Session) -> WebsiteAudit:
    try:
        response = session.get(
            "https://www.googleapis.com/pagespeedonline/v5/runPagespeed",
            params={
                "url": audit.normalized_url,
                "strategy": "mobile",
                "category": "performance",
                "key": api_key,
            },
            headers={"User-Agent": USER_AGENT},
            timeout=20,
        )
        response.raise_for_status()
        return apply_pagespeed_result(audit, response.json())
    except (requests.RequestException, ValueError, TypeError):
        return audit


def audit_website(
    value: str,
    session: requests.Session | None = None,
    timeout: int = 7,
    pagespeed_key: str = "",
) -> WebsiteAudit:
    audit = normalize_website(value)
    if audit.state != "unknown":
        return audit

    client = session or requests.Session()
    raw = value.strip()
    urls = [audit.normalized_url]
    if "://" not in raw:
        urls.append(f"http://{raw}")

    saw_timeout = False
    last_status: int | None = None
    for url in urls:
        try:
            response = client.get(
                url,
                headers={"User-Agent": USER_AGENT},
                timeout=timeout,
                allow_redirects=True,
            )
        except requests.Timeout:
            saw_timeout = True
            continue
        except requests.RequestException:
            continue

        last_status = response.status_code
        final_url = getattr(response, "url", url) or url
        if response.status_code in (401, 403, 429):
            return WebsiteAudit(
                state="limited",
                normalized_url=final_url,
                https=final_url.startswith("https://"),
                status=response.status_code,
            )
        if 200 <= response.status_code < 400:
            viewport, title, description = _parse_html_signals(response.text)
            result = WebsiteAudit(
                state="reachable",
                normalized_url=final_url,
                https=final_url.startswith("https://"),
                status=response.status_code,
                mobile_viewport=viewport,
                title_present=title,
                description_present=description,
            )
            return _check_pagespeed(result, pagespeed_key, client) if pagespeed_key else result

    if saw_timeout:
        return WebsiteAudit(state="unknown", normalized_url=audit.normalized_url, error="Таймаут")
    return WebsiteAudit(state="broken", normalized_url=audit.normalized_url, status=last_status)


def score_lead(lead: Lead, audit: WebsiteAudit) -> tuple[int, list[str]]:
    need_score = 0
    reasons: list[str] = []

    if audit.state == "missing":
        need_score = 60
        reasons.append("сайт не указан в источнике")
    elif audit.state == "social":
        need_score = 45
        reasons.append("вместо сайта соцсеть")
    elif audit.state == "broken":
        need_score = 50
        reasons.append("сайт не открывается")
    elif audit.state == "unknown":
        reasons.append("сайт не удалось проверить")
    elif audit.state in ("reachable", "limited"):
        if not audit.https:
            need_score += 15
            reasons.append("нет HTTPS")
        if audit.state == "limited":
            reasons.append("аудит сайта ограничен")
        else:
            if audit.mobile_viewport is False:
                need_score += 15
                reasons.append("нет мобильной адаптации")
            if audit.title_present is False:
                need_score += 5
                reasons.append("нет title")
            if audit.description_present is False:
                need_score += 5
                reasons.append("нет description")
            if audit.mobile_score is not None and audit.mobile_score < 50:
                need_score += 10
                reasons.append("низкая мобильная скорость")
            elif audit.mobile_score is not None and audit.mobile_score < 70:
                need_score += 5
                reasons.append("средняя мобильная скорость")

    contact_score = 0
    if lead.phone:
        contact_score += 20
        reasons.append("есть телефон")
    if lead.email:
        contact_score += 10
        reasons.append("есть email")
    if lead.social:
        contact_score += 5
        reasons.append("есть соцсеть/мессенджер")

    return min(100, min(70, need_score) + min(30, contact_score)), reasons


def _confirmed_problem(audit: WebsiteAudit) -> str:
    if audit.state == "missing":
        return "в открытой карточке не указан сайт"
    if audit.state == "social":
        return "вместо сайта указана только страница в соцсети"
    if audit.state == "broken":
        return "указанный сайт сейчас не открывается"
    if audit.state == "reachable":
        if not audit.https:
            return "сайт работает без защищённого HTTPS"
        if audit.mobile_viewport is False:
            return "сайт не адаптирован для мобильных экранов"
        if audit.title_present is False or audit.description_present is False:
            return "на сайте не заполнены базовые SEO-метаданные"
        if audit.mobile_score is not None and audit.mobile_score < 70:
            return "мобильная версия сайта загружается медленно"
    return "сайт можно усилить для получения обращений"


def render_outreach(lead: Lead, audit: WebsiteAudit, channel: str) -> str:
    problem = _confirmed_problem(audit)
    location = f" в городе {lead.city}" if lead.city else ""
    if channel == "message":
        return (
            f"Здравствуйте! Нашёл компанию «{lead.name}» при поиске местных услуг{location}. "
            f"Заметил, что {problem}. Могу предложить простой вариант сайта или точечную доработку, "
            "чтобы клиентам было легче увидеть услуги и оставить заявку. Могу прислать короткий план?"
        )
    if channel == "email":
        return (
            f"Тема: Идея по сайту для «{lead.name}»\n\n"
            f"Здравствуйте! Нашёл вашу компанию{location} и заметил, что {problem}. "
            "Я создаю и дорабатываю сайты для локального бизнеса. Могу бесплатно прислать короткий "
            "список улучшений и оценку объёма работ. Если актуально, ответьте на это письмо."
        )
    if channel == "call":
        return (
            f"Начало: Здравствуйте! Подскажите, я попал в компанию «{lead.name}»? "
            f"Я посмотрел, как вы представлены в интернете, и заметил, что {problem}.\n"
            "Вопрос: Вы сейчас рассматриваете создание сайта или доработку текущего?\n"
            "Предложение: Я могу сначала прислать короткий разбор без обязательств, а после него вы решите, нужно ли продолжать."
        )
    raise ValueError("Неизвестный канал обращения.")


def build_overpass_query(
    preset: str,
    keyword: str,
    bbox: tuple[float, float, float, float],
    limit: int,
) -> str:
    if not 1 <= limit <= 200:
        raise ValueError("Лимит должен быть от 1 до 200.")
    south, west, north, east = bbox
    box = f"{south},{west},{north},{east}"
    escaped = re.escape(keyword.strip())
    keyword_filter = f'["name"~"{escaped}",i]' if escaped else ""

    selectors: list[str] = []
    if preset in PRESETS:
        for key, values in PRESETS[preset]:
            tag_filter = f'["{key}"~"^({values})$"]' if "|" in values else f'["{key}"="{values}"]'
            selectors.append(f"nwr({box}){tag_filter}{keyword_filter};")
    elif preset == "Другая ниша" and escaped:
        for key in ("shop", "craft", "office", "amenity", "healthcare"):
            selectors.append(f'nwr({box})["{key}"]["name"~"{escaped}",i];')
    else:
        raise ValueError("Выберите пресет или укажите ключевое слово для другой ниши.")

    return "[out:json][timeout:25];\n(\n  " + "\n  ".join(selectors) + f"\n);\nout center {limit} tags;"


def _first(tags: dict[str, Any], *names: str) -> str:
    for name in names:
        value = tags.get(name)
        if value:
            return str(value).strip()
    return ""


def _normalize_social(tags: dict[str, Any]) -> str:
    telegram = _first(tags, "contact:telegram", "telegram")
    if telegram:
        return telegram if "://" in telegram else f"https://t.me/{telegram.lstrip('@')}"
    vk = _first(tags, "contact:vk", "vk")
    if vk:
        return vk if "://" in vk else f"https://vk.com/{vk.lstrip('@')}"
    return _first(tags, "contact:instagram", "instagram", "contact:facebook", "facebook")


def parse_osm_elements(payload: dict[str, Any], city: str) -> list[Lead]:
    leads: list[Lead] = []
    for element in payload.get("elements", []):
        tags = element.get("tags") or {}
        name = _first(tags, "name", "brand", "operator")
        if not name:
            continue
        street = _first(tags, "addr:street", "addr:place")
        house = _first(tags, "addr:housenumber")
        address = ", ".join(part for part in (street, house) if part) or _first(tags, "addr:full")
        category = _first(tags, "craft", "shop", "office", "amenity", "healthcare")
        element_type = str(element.get("type", "node"))
        element_id = str(element.get("id", ""))
        leads.append(
            Lead(
                lead_key=f"osm:{element_type}:{element_id}",
                name=name,
                category=category,
                city=city,
                address=address,
                phone=_first(tags, "contact:phone", "phone"),
                email=_first(tags, "contact:email", "email"),
                social=_normalize_social(tags),
                website=_first(tags, "contact:website", "website", "url"),
                source="OpenStreetMap",
                source_url=f"https://www.openstreetmap.org/{element_type}/{element_id}",
            )
        )
    return leads


def resolve_city_bbox(city: str, session: requests.Session | None = None) -> tuple[float, float, float, float]:
    client = session or requests.Session()
    response = client.get(
        "https://nominatim.openstreetmap.org/search",
        params={
            "q": f"{city}, Россия",
            "format": "jsonv2",
            "limit": 1,
            "countrycodes": "ru",
        },
        headers={"User-Agent": USER_AGENT},
        timeout=15,
    )
    response.raise_for_status()
    payload = response.json()
    if not payload:
        raise ValueError(f"Город «{city}» не найден.")
    south, north, west, east = payload[0]["boundingbox"]
    return float(south), float(west), float(north), float(east)


def collect_osm(
    city: str,
    preset: str,
    keyword: str,
    limit: int = 50,
    only_with_contacts: bool = True,
    session: requests.Session | None = None,
) -> list[Lead]:
    client = session or requests.Session()
    try:
        bbox = resolve_city_bbox(city, client)
    except requests.RequestException as error:
        raise RuntimeError(f"Не удалось определить границы города: {error}") from error
    query = build_overpass_query(preset, keyword, bbox, limit)
    last_error: Exception | None = None
    payload: dict[str, Any] | None = None

    for endpoint in OVERPASS_ENDPOINTS:
        try:
            response = client.post(
                endpoint,
                data={"data": query},
                headers={"User-Agent": USER_AGENT},
                timeout=35,
            )
            response.raise_for_status()
            payload = response.json()
            break
        except (requests.RequestException, ValueError) as error:
            last_error = error

    if payload is None:
        raise RuntimeError(f"Источники OpenStreetMap недоступны: {last_error}")

    leads = parse_osm_elements(payload, city)
    if only_with_contacts:
        leads = [lead for lead in leads if lead.phone or lead.email or lead.social or lead.website]
    return leads[:limit]


def enrich_leads(
    leads: list[Lead],
    pagespeed_key: str = "",
    audit_func: Any = audit_website,
) -> list[Lead]:
    def enrich(lead: Lead) -> Lead:
        audit = audit_func(lead.website or lead.social, pagespeed_key=pagespeed_key)
        lead.audit = audit
        if lead.website and audit.normalized_url:
            lead.website = audit.normalized_url
        lead.score, lead.reasons = score_lead(lead, audit)
        return lead

    with ThreadPoolExecutor(max_workers=min(5, max(1, len(leads)))) as executor:
        enriched = list(executor.map(enrich, leads))
    return sorted(enriched, key=lambda item: item.score, reverse=True)


def dry_run_leads() -> list[Lead]:
    samples = [
        (Lead(name="Мастер окон", lead_key="dry:1", category="Окна", city="Екатеринбург", phone="+7 343 000-00-01", source="Dry Run"), WebsiteAudit(state="missing")),
        (Lead(name="Сантехник рядом", lead_key="dry:2", category="Сантехники", city="Екатеринбург", phone="+7 343 000-00-02", social="https://vk.com/santeh", website="https://vk.com/santeh", source="Dry Run"), WebsiteAudit(state="social", normalized_url="https://vk.com/santeh", https=True)),
        (Lead(name="Урал-Сервис", lead_key="dry:3", category="Ремонт", city="Екатеринбург", email="info@example.ru", website="http://example.ru", source="Dry Run"), WebsiteAudit(state="reachable", normalized_url="http://example.ru", https=False, mobile_viewport=False, title_present=True, description_present=False)),
    ]
    result: list[Lead] = []
    for lead, audit in samples:
        lead.audit = audit
        lead.score, lead.reasons = score_lead(lead, audit)
        result.append(lead)
    return sorted(result, key=lambda item: item.score, reverse=True)


def clean_cell(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _fallback_key(values: dict[str, str]) -> str:
    contact = values.get("phone") or values.get("email")
    if contact:
        token = re.sub(r"\W+", "", contact.lower())
    else:
        token = re.sub(r"\W+", "", f"{values.get('name', '')}:{values.get('address', '')}".lower())
    return f"import:{token}"


def lead_from_row(row: dict[str, Any]) -> Lead | None:
    values = {field: clean_cell(row.get(field)) for field in INPUT_FIELDS}
    if not values["name"]:
        return None
    return Lead(lead_key=_fallback_key(values), **values)


def import_csv(path_or_file: str | io.BytesIO) -> list[Lead]:
    if isinstance(path_or_file, str):
        handle = open(path_or_file, newline="", encoding="utf-8-sig")
        close = True
    else:
        handle = io.StringIO(path_or_file.getvalue().decode("utf-8-sig"))
        close = False
    try:
        reader = csv.DictReader(handle)
        if not reader.fieldnames or "name" not in reader.fieldnames:
            raise ValueError("Во входном CSV обязательна колонка name.")
        return [lead for row in reader if (lead := lead_from_row(row))]
    finally:
        if close:
            handle.close()


def import_xlsx(path_or_file: str | io.BytesIO) -> list[Lead]:
    workbook = load_workbook(path_or_file, read_only=True, data_only=True)
    try:
        rows = workbook.worksheets[0].iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            raise ValueError("Во входном XLSX обязательна колонка name.")
        fieldnames = [clean_cell(header) for header in headers]
        if "name" not in fieldnames:
            raise ValueError("Во входном XLSX обязательна колонка name.")
        result = []
        for values in rows:
            row = {fieldnames[index]: value for index, value in enumerate(values) if index < len(fieldnames)}
            lead = lead_from_row(row)
            if lead:
                result.append(lead)
        return result
    finally:
        workbook.close()


def lead_to_export_row(lead: Lead) -> dict[str, Any]:
    return {
        "score": lead.score,
        "priority": lead.priority,
        "reasons": "; ".join(lead.reasons),
        **{field: getattr(lead, field) for field in EXPORT_FIELDS[3:]},
    }


def export_csv_bytes(leads: list[Lead]) -> bytes:
    output = io.StringIO(newline="")
    writer = csv.DictWriter(output, fieldnames=EXPORT_FIELDS)
    writer.writeheader()
    for lead in sorted(leads, key=lambda item: item.score, reverse=True):
        writer.writerow(lead_to_export_row(lead))
    return output.getvalue().encode("utf-8-sig")


def export_xlsx_bytes(leads: list[Lead]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Лиды"
    sheet.append(EXPORT_FIELDS)
    for lead in sorted(leads, key=lambda item: item.score, reverse=True):
        row = lead_to_export_row(lead)
        sheet.append([row[field] for field in EXPORT_FIELDS])
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def export_csv(leads: list[Lead], output: str) -> None:
    with open(output, "wb") as handle:
        handle.write(export_csv_bytes(leads))


def configure_logging(path: str = "lead_finder.log") -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s: %(message)s",
        handlers=[logging.FileHandler(path, encoding="utf-8")],
        force=True,
    )
