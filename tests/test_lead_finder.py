import csv
import io
import os
import tempfile
import unittest
from zipfile import BadZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from lead_finder import (
    MAPPING_FIELDS,
    Lead,
    WebsiteAudit,
    apply_pagespeed_result,
    export_csv_bytes,
    export_xlsx_bytes,
    guess_mapping,
    import_csv,
    import_xlsx,
    lead_from_row,
    read_headers,
)


class LeadFinderFileTests(unittest.TestCase):
    def test_pagespeed_result_sets_mobile_score(self):
        audit = WebsiteAudit(state="reachable", normalized_url="https://slow.ru")

        result = apply_pagespeed_result(
            audit,
            {"lighthouseResult": {"categories": {"performance": {"score": 0.42}}}},
        )

        self.assertEqual(result.mobile_score, 42)

    def test_import_csv_reads_known_columns_and_builds_stable_key(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "manual.csv")
            with open(path, "w", newline="", encoding="utf-8") as handle:
                writer = csv.DictWriter(handle, fieldnames=["name", "phone", "website"])
                writer.writeheader()
                writer.writerow({"name": "Мастер окон", "phone": "+7 343", "website": ""})

            leads = import_csv(path)

        self.assertEqual(len(leads), 1)
        self.assertEqual(leads[0].name, "Мастер окон")
        self.assertTrue(leads[0].lead_key.startswith("import:"))

    def test_import_csv_requires_name_column(self):
        data = io.BytesIO(b"website\nexample.ru\n")

        with self.assertRaises(ValueError):
            import_csv(data)

    def test_import_xlsx_reads_first_sheet(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "manual.xlsx")
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["name", "category", "phone"])
            sheet.append(["Сантехник", "plumber", "+7"])
            workbook.save(path)
            workbook.close()

            leads = import_xlsx(path)

        self.assertEqual(len(leads), 1)
        self.assertEqual(leads[0].category, "plumber")

    def test_import_xlsx_requires_name_column(self):
        data = io.BytesIO()
        workbook = Workbook()
        workbook.active.append(["website"])
        workbook.active.append(["example.ru"])
        workbook.save(data)
        workbook.close()
        data.seek(0)

        with self.assertRaises(ValueError):
            import_xlsx(data)

    def test_import_csv_maps_localized_headers_and_joins_repeated_columns(self):
        data = io.BytesIO(
            "Наименование,Адрес,Телефон 1,Телефон 2,Веб-сайт,Широта,Долгота\r\n"
            "Мастер окон,Ленина 1,+7 343 000-00-01,+7 343 000-00-02,example.ru,56.83,60.6\r\n".encode("utf-8")
        )

        leads = import_csv(data)

        self.assertEqual(len(leads), 1)
        self.assertEqual(leads[0].name, "Мастер окон")
        self.assertEqual(leads[0].address, "Ленина 1")
        self.assertEqual(leads[0].phone, "+7 343 000-00-01, +7 343 000-00-02")
        self.assertEqual(leads[0].website, "example.ru")
        self.assertEqual(leads[0].verification_status, "source_provided")
        self.assertEqual(leads[0].latitude, 56.83)
        self.assertEqual(leads[0].longitude, 60.6)

    def test_import_csv_manual_mapping_overrides_guess(self):
        data = io.BytesIO("Наименование,Комментарий\r\nМастер окон,Сантехник рядом\r\n".encode("utf-8"))

        leads = import_csv(data, {"name": ["Комментарий"]})

        self.assertEqual(len(leads), 1)
        self.assertEqual(leads[0].name, "Сантехник рядом")
        self.assertEqual(leads[0].address, "")

    def test_import_csv_without_name_source_raises(self):
        data = io.BytesIO("Комментарий,Примечание\r\nтекст,текст\r\n".encode("utf-8"))

        with self.assertRaises(ValueError):
            import_csv(data)

    def test_import_csv_broken_coordinate_becomes_none(self):
        data = io.BytesIO("Наименование,Широта,Долгота\r\nМастер окон,нет данных,-\r\n".encode("utf-8"))

        leads = import_csv(data)

        self.assertIsNone(leads[0].latitude)
        self.assertIsNone(leads[0].longitude)

    def test_import_xlsx_maps_localized_headers(self):
        data = io.BytesIO()
        workbook = Workbook()
        workbook.active.append(["Наименование", "Рубрики", "E-mail 1"])
        workbook.active.append(["Сантехник рядом", "Сантехники", "info@example.ru"])
        workbook.save(data)
        workbook.close()
        data.seek(0)

        leads = import_xlsx(data)

        self.assertEqual(len(leads), 1)
        self.assertEqual(leads[0].category, "Сантехники")
        self.assertEqual(leads[0].email, "info@example.ru")

    def test_import_csv_joins_columns_with_identical_headers(self):
        data = io.BytesIO(
            "Наименование,Телефон,Телефон\r\nМастер окон,+7 343 000-00-01,+7 343 000-00-02\r\n".encode("utf-8")
        )

        leads = import_csv(data)

        self.assertEqual(leads[0].phone, "+7 343 000-00-01, +7 343 000-00-02")

    def test_import_csv_keeps_single_email_when_columns_repeat(self):
        data = io.BytesIO(
            "Наименование,E-mail 1,E-mail 2\r\nМастер окон,first@example.ru,second@example.ru\r\n".encode("utf-8")
        )

        leads = import_csv(data)

        self.assertEqual(leads[0].email, "first@example.ru")

    def test_import_xlsx_joins_columns_with_identical_headers(self):
        data = io.BytesIO()
        workbook = Workbook()
        workbook.active.append(["Наименование", "Телефон", "Телефон"])
        workbook.active.append(["Мастер окон", "+7 343 000-00-01", "+7 343 000-00-02"])
        workbook.save(data)
        workbook.close()
        data.seek(0)

        leads = import_xlsx(data)

        self.assertEqual(leads[0].phone, "+7 343 000-00-01, +7 343 000-00-02")

    def test_read_headers_numbers_repeated_columns(self):
        data = io.BytesIO("Наименование,Телефон,Телефон\r\n".encode("utf-8"))

        self.assertEqual(read_headers(data, "csv"), ["Наименование", "Телефон", "Телефон (2)"])

    def test_repeated_headers_never_collide_with_generated_names(self):
        data = io.BytesIO("Наименование,Телефон,Телефон (2),Телефон\r\n".encode("utf-8"))

        headers = read_headers(data, "csv")

        self.assertEqual(headers, ["Наименование", "Телефон", "Телефон (2)", "Телефон (3)"])
        self.assertEqual(len(set(headers)), len(headers))

    def test_import_csv_keeps_every_value_of_three_repeated_columns(self):
        data = io.BytesIO(
            "Наименование,Телефон,Телефон (2),Телефон\r\nМастер окон,+7 001,+7 002,+7 003\r\n".encode("utf-8")
        )

        leads = import_csv(data)

        self.assertEqual(leads[0].phone, "+7 001, +7 002, +7 003")

    def test_columns_without_headers_get_distinct_names(self):
        data = io.BytesIO("Наименование,,Телефон,\r\nМастер окон,первое,+7 001,второе\r\n".encode("utf-8"))

        headers = read_headers(data, "csv")
        leads = import_csv(data, {"name": ["Наименование"], "address": headers[1:2], "city": headers[3:4]})

        self.assertEqual(headers, ["Наименование", "Колонка 2", "Телефон", "Колонка 4"])
        self.assertEqual(len(set(headers)), len(headers))
        self.assertEqual(leads[0].address, "первое")
        self.assertEqual(leads[0].city, "второе")

    def test_broken_extension_raises_recognized_error(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "leads.xls")
            with open(path, "w", encoding="utf-8") as handle:
                handle.write("name,phone\n")

            with self.assertRaises(InvalidFileException):
                read_headers(path, "xlsx")
            with self.assertRaises(InvalidFileException):
                import_xlsx(path)

    def test_broken_xlsx_raises_recognized_error(self):
        data = io.BytesIO("name,phone\r\nМастер окон,+7 343\r\n".encode("utf-8"))

        with self.assertRaises(BadZipFile):
            read_headers(data, "xlsx")
        with self.assertRaises(BadZipFile):
            import_xlsx(data)

    def test_guess_mapping_recognizes_every_field_by_its_own_name(self):
        for field_name in MAPPING_FIELDS:
            with self.subTest(field=field_name):
                mapping = guess_mapping([field_name.upper()])

                self.assertEqual(mapping.get(field_name), [field_name.upper()])

    def test_lead_from_row_without_mapping_reads_exact_fields(self):
        lead = lead_from_row({"name": "Мастер окон", "phone": "+7 343", "website": "example.ru"})

        self.assertEqual(lead.name, "Мастер окон")
        self.assertEqual(lead.website, "example.ru")
        self.assertEqual(lead.verification_status, "source_provided")
        self.assertIsNone(lead.latitude)

    def test_guess_mapping_prefers_exact_field_name(self):
        mapping = guess_mapping(["name", "Наименование", "Сайт", "website"])

        self.assertEqual(mapping["name"], ["name", "Наименование"])
        self.assertEqual(mapping["website"], ["Сайт", "website"])

    def test_read_headers_rewinds_stream_for_later_import(self):
        data = io.BytesIO("Наименование,Телефон\r\nМастер окон,+7 343\r\n".encode("utf-8"))

        headers = read_headers(data, "csv")
        leads = import_csv(data)

        self.assertEqual(headers, ["Наименование", "Телефон"])
        self.assertEqual(len(leads), 1)
        self.assertEqual(leads[0].phone, "+7 343")

    def test_read_headers_rewinds_xlsx_stream_for_later_import(self):
        data = io.BytesIO()
        workbook = Workbook()
        workbook.active.append(["Наименование", "Телефон"])
        workbook.active.append(["Мастер окон", "+7 343"])
        workbook.save(data)
        workbook.close()
        data.seek(0)

        headers = read_headers(data, "xlsx")
        leads = import_xlsx(data)

        self.assertEqual(headers, ["Наименование", "Телефон"])
        self.assertEqual(len(leads), 1)

    def test_csv_export_sorts_by_score_and_contains_reasons(self):
        leads = [
            Lead(name="Низкий", lead_key="1", score=10, reasons=["есть телефон"]),
            Lead(name="Высокий", lead_key="2", score=80, reasons=["сайт не указан в источнике"]),
        ]

        rows = list(csv.DictReader(io.StringIO(export_csv_bytes(leads).decode("utf-8-sig"))))

        self.assertEqual(rows[0]["name"], "Высокий")
        self.assertEqual(rows[0]["reasons"], "сайт не указан в источнике")

    def test_xlsx_export_contains_rows(self):
        data = export_xlsx_bytes([Lead(name="Компания", lead_key="1", score=70)])
        workbook = load_workbook(io.BytesIO(data), read_only=True)
        try:
            rows = list(workbook.active.iter_rows(values_only=True))
        finally:
            workbook.close()

        self.assertEqual(rows[0][0], "score")
        self.assertEqual(rows[1][3], "Компания")


if __name__ == "__main__":
    unittest.main()
